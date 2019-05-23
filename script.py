import openpyxl,datetime,bs4

today = datetime.date.today()
while today.day >= 2:
  today += datetime.timedelta(days= -1)  #今月の１日を取得

wb = openpyxl.load_workbook('analysis.xlsx')
sheet = wb.get_sheet_by_name('monthly_data')




aray = []

for i in range(1,sheet.max_row+1):
  aray[i] = sheet.cell(row=i,column=1).value

print(aray)
